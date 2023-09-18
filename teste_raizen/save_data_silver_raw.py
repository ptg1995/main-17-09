import openpyxl

def save_tables_silver_raw(path, sheets:dict):
    # path = "/root/airflow/dags/bronze_raw/data_full.xlsx"
    # sheets = {"DPCache_m3": "oil_derivative", "DPCache_m3_2": "diesel"}
    workbook = openpyxl.load_workbook(path)
    for sheet_name, file_name in sheets.items():
        sheet = workbook[sheet_name]
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.active
        new_sheet.title = sheet_name

        for row in sheet.iter_rows(values_only=True):
            new_sheet.append(row)

        new_path = f"/root/airflow/dags/silver_raw/{file_name}_data.xlsx"
        new_workbook.save(new_path)
    print("Sheets salvos com sucesso!")